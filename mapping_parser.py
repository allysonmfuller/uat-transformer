"""
mapping_parser.py
Reads the strategy mapping template Excel file and extracts
per-field data (Andar location/logic, DTracker location/logic,
SCRM label) keyed by SCRM API name.
"""

import re
from io import BytesIO

import openpyxl
import pandas as pd


# ── Keyword lists for detecting each row in the mapping sheet ──────────────────
ROW_KEYWORDS = {
    "api_name": [
        "npsp api name", "api name", "api names",
        "salesforce api", "scrm api",
    ],
    "scrm_label": [
        "andar label", "andar label:",
        "dt label", "dtracker label", "scrm label",
        "field label",
        # NOTE: "andar ui label" is intentionally excluded here —
        # it belongs in andar_field (Andar Location), not SCRM Field Label
    ],
    # Primary Andar field row (human-readable path OR import header when only one exists)
    "andar_field": [
        "^andar field$",
        "andar field:",
        "andar field - individuals",   # Address-style: this IS the field row
        "andar field - organizations",
        "andar field\n", "andar field (import",
        "andar location",
        "andar ui label",              # Notes Task-style templates
    ],
    # Secondary Andar row — ONLY matches "Andar API Field / Import Header"
    # (a separate row that appears AFTER the primary andar_field row)
    # Combined with andar_field in Andar Location cell
    "andar_import": [
        "andar api field /",           # "Andar API Field / Import Header"
        "andar api field/",
        "^andar api field$",
    ],
    "andar_logic": [
        "andar logic", "andar logic:",
    ],
    "dt_field": [
        "dtracker field", "dt field", "dt api field",
        "dtracker field:", "dt field - individuals",
        "dt field - organizations",
        "dt api field /import header",
        "dt api field/import header",
    ],
    # Secondary DT row — import header
    "dt_import": [
        "dt api field /",
        "dt api field/",
        "dtracker api field /",
    ],
    "dt_logic": [
        "dtracker logic", "dt logic", "dtracker logic:",
    ],
}


def _match(label: str, keywords: list) -> bool:
    """Return True if label matches any keyword (case-insensitive)."""
    label_lower = label.strip().lower()
    for kw in keywords:
        if kw.startswith("^") and kw.endswith("$"):
            if re.fullmatch(kw[1:-1], label_lower):
                return True
        elif kw in label_lower:
            return True
    return False


def _find_mapping_sheet(wb) -> str:
    """Return the name of the mapping sheet (not Legend)."""
    for name in wb.sheetnames:
        if name.lower() not in ("legend",):
            return name
    return wb.sheetnames[0]


def _find_rows(df) -> dict:
    """
    Scan the mapping sheet and identify key row indices.

    Handles:
    - Named rows matched by keyword
    - Bare "Logic" / "LOGIC" rows assigned by position
      (first after andar_field = andar_logic, first after dt_field = dt_logic)
    """
    row_map = {k: None for k in ROW_KEYWORDS}
    generic_logic_rows = []

    for i, row in df.iterrows():
        label = str(row.iloc[0]).strip()
        if label in ("nan", ""):
            continue

        label_lower = label.lower()

        # Bare "Logic" or "LOGIC" — position-dependent
        if re.fullmatch(r"logic", label_lower) or re.fullmatch(r"notes", label_lower):
            generic_logic_rows.append(i)
            continue

        for key, keywords in ROW_KEYWORDS.items():
            if row_map[key] is None and _match(label, keywords):
                row_map[key] = i
                break

    # Assign bare Logic rows by position
    andar_field_row = row_map["andar_field"] or 0
    dt_field_row    = row_map["dt_field"]    or 0

    for logic_row in generic_logic_rows:
        if row_map["andar_logic"] is None and logic_row > andar_field_row:
            row_map["andar_logic"] = logic_row
        elif row_map["dt_logic"] is None and logic_row > dt_field_row:
            row_map["dt_logic"] = logic_row

    return row_map


def _infer_object_name(filename: str) -> str:
    """
    Extract a human-readable object name from the template filename.
    Add new objects to the `known` dict as needed.
    Order matters — more specific keys must come before general ones.
    """
    known = {
        # ── Object Set 1-3 ────────────────────────────────────────────────────
        "opppayment":              "Payment",
        "oppayment":               "Payment",
        "payment":                 "Payment",
        "allocation":              "Allocation",
        "address":                 "Address",
        "affiliation":             "Affiliation",
        "accountrelationship":     "Account Relationship",
        "individualrelationship":  "Individual Relationship",
        "generalaccountingunit":   "GAU",
        "gau":                     "GAU",
        "contact":                 "Contact",
        "opportunity":             "Opportunity",
        "account":                 "Account",
        # ── Object Set 4 ─────────────────────────────────────────────────────
        "campaignmember":          "Campaign Member",
        "campaign_member":         "Campaign Member",
        "volunteer_job":           "Volunteer Job",
        "volunteerjob":            "Volunteer Job",
        "volunteer_hours":         "Volunteer Hours",
        "volunteerhours":          "Volunteer Hours",
        "teams__c":                "Teams",
        "volunteerteammember":     "Volunteer Team Member",
        "note_and_communication":  "Note and Communication Code",
        "noteandcommunication":    "Note and Communication Code",
        "communication_log":       "Communication Log Task",
        "communicationlog":        "Communication Log Task",
        "attachment":              "Attachment",
        "notes_task":              "Notes Task",
        "notestask":               "Notes Task",
        "engagement_plan_task":    "Engagement Plan Task",
        "engagementplantask":      "Engagement Plan Task",
        "engagement_plan__c":      "Engagement Plan",
        "engagementplan":          "Engagement Plan",
        # ── Generic fallbacks — keep at the bottom ────────────────────────────
        "campaign":                "Campaign",
    }

    lower = filename.lower()
    for key, label in known.items():
        if key in lower:
            return label

    match = re.search(r"__([A-Za-z]+)__c", filename)
    if match:
        raw = match.group(1)
        spaced = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", raw)
        return spaced.strip()

    return "Object"


def parse_mapping_template(template_bytes: bytes, filename: str = "") -> dict:
    """
    Parse the strategy mapping template.

    Andar Location is built by combining:
      - andar_field row  (human-readable path, e.g. "Volunteer Opportunity / Assigned")
      - andar_import row (import header/API field, e.g. "addresses.addresstype")
    If both exist for a field, they are joined with a newline.

    Returns:
        {
          "fields": {api_name: {scrm_label, andar_field, andar_logic,
                                dt_field, dt_logic}},
          "meta":   {object_name, api_names_ordered, has_scrm_labels}
        }
    """
    wb         = openpyxl.load_workbook(BytesIO(template_bytes), data_only=True)
    sheet_name = _find_mapping_sheet(wb)
    df         = pd.read_excel(
        BytesIO(template_bytes), sheet_name=sheet_name, header=None
    )

    row_map = _find_rows(df)

    if row_map["api_name"] is None:
        raise ValueError(
            "Could not find an API Name row in the mapping template. "
            "Expected a row labelled 'NPSP API Name' or 'API Names'."
        )

    # ── Build col → API name index ─────────────────────────────────────────────
    api_row   = df.iloc[row_map["api_name"]].tolist()
    api_clean = []
    for v in api_row:
        s = str(v).strip().split("\n")[0].strip()
        api_clean.append(s if s != "nan" else "")

    api_names_ordered = [a for a in api_clean[1:] if a]

    def row_to_dict(row_idx):
        if row_idx is None:
            return {}
        row = df.iloc[row_idx].tolist()
        return {
            api_clean[i]: str(v).strip()
            for i, v in enumerate(row)
            if i > 0 and api_clean[i] and str(v).strip() not in ("nan", "")
        }

    scrm_labels    = row_to_dict(row_map["scrm_label"])
    andar_fields   = row_to_dict(row_map["andar_field"])
    andar_imports  = row_to_dict(row_map["andar_import"])
    andar_logics   = row_to_dict(row_map["andar_logic"])
    dt_fields      = row_to_dict(row_map["dt_field"])
    dt_imports     = row_to_dict(row_map["dt_import"])
    dt_logics      = row_to_dict(row_map["dt_logic"])

    # ── Merge per-field dicts ──────────────────────────────────────────────────
    fields = {}
    for api in api_names_ordered:
        # Combine Andar field path + import header into one Andar Location value
        andar_path   = andar_fields.get(api, "")
        andar_import = andar_imports.get(api, "")
        if andar_path and andar_import:
            andar_combined = f"{andar_path}\n{andar_import}"
        else:
            andar_combined = andar_path or andar_import

        # Same for DT
        dt_path   = dt_fields.get(api, "")
        dt_import = dt_imports.get(api, "")
        if dt_path and dt_import:
            dt_combined = f"{dt_path}\n{dt_import}"
        else:
            dt_combined = dt_path or dt_import

        fields[api] = {
            "scrm_label":  scrm_labels.get(api, ""),
            "andar_field": andar_combined,
            "andar_logic": andar_logics.get(api, ""),
            "dt_field":    dt_combined,
            "dt_logic":    dt_logics.get(api, ""),
        }

    object_name = _infer_object_name(filename)

    return {
        "fields": fields,
        "meta": {
            "object_name":       object_name,
            "api_names_ordered": api_names_ordered,
            "has_scrm_labels":   bool(scrm_labels),
        },
    }
