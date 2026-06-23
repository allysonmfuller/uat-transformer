"""
styles.py
All formatting constants and cell-styling helpers.
One place to change fonts, borders, or colours across the whole app.
"""

from copy import copy

from openpyxl.styles import Alignment, Border, Font, Side

# ── Constants ──────────────────────────────────────────────────────────────────
FONT_NAME = "Calibri"
FONT_SIZE = 12

THIN      = Side(style="thin")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NO_BORDER   = Border()


# ── Helpers ────────────────────────────────────────────────────────────────────

def std_font(bold: bool = False) -> Font:
    """Return a standard Calibri 12 font, optionally bold."""
    return Font(name=FONT_NAME, size=FONT_SIZE, bold=bold)


def apply_std(cell, bold: bool = False, wrap: bool = False, border: bool = True) -> None:
    """
    Apply standard formatting to a cell:
    - Calibri 12, optional bold
    - Optional wrap text
    - Thin border on all four sides (or no border if border=False)
    """
    cell.font      = std_font(bold=bold)
    cell.alignment = Alignment(wrap_text=wrap)
    cell.border    = copy(THIN_BORDER) if border else NO_BORDER


def copy_style(src, dst) -> None:
    """Copy all style attributes from one cell to another."""
    if src.has_style:
        dst.font         = copy(src.font)
        dst.fill         = copy(src.fill)
        dst.border       = copy(src.border)
        dst.alignment    = copy(src.alignment)
        dst.number_format = src.number_format


def swap_cols(ws, col_a: int, col_b: int, min_row: int = 1) -> None:
    """
    Swap values AND styles between two column indices across all rows.
    col_a, col_b are 1-based column numbers.
    """
    for r in range(min_row, ws.max_row + 1):
        ca, cb = ws.cell(row=r, column=col_a), ws.cell(row=r, column=col_b)
        ca.value, cb.value = cb.value, ca.value
        for attr in ("font", "fill", "border", "alignment"):
            a_val = copy(getattr(ca, attr))
            b_val = copy(getattr(cb, attr))
            setattr(ca, attr, b_val)
            setattr(cb, attr, a_val)
