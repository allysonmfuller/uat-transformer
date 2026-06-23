"""
template_builder.py
Builds the output Excel workbook from scratch using:
  - The hardcoded blank UAT template (embedded as bytes in this file)
  - The parsed mapping data from mapping_parser.py
  - The user's source system selections

This replaces the need to upload a UAT tracking file.
"""

from copy import copy
from io import BytesIO

import openpyxl
from openpyxl.styles import PatternFill

from styles import THIN_BORDER, NO_BORDER, apply_std, std_font
from openpyxl.styles import Alignment, Border

# ── Branch sheet config ────────────────────────────────────────────────────────
# Maps source system name → list of (sheet_name, branch_type)
BRANCH_CONFIGS = {
    "Andar":         [("WHALIF", "andar"), ("WPEI",   "andar")],
    "DTracker":      [("WFREDE", "dt"),    ("WSTJOH", "dt")],
    "Raiser's Edge": [("WRE1",   "re"),    ("WRE2",   "re")],
}

# ── Column layout (1-based) matches blank template exactly ────────────────────
# A=DB1/Andar, B=DB2/DT, C=DB3/RE, D=Mapping Notes,
# E=SCRM Field Name, F=SCRM Field Label, G+=UAT cols
COL_ANDAR     = 1   # Andar Location
COL_DT        = 2   # DT Location
COL_RE        = 3   # RE Location (placeholder)
COL_NOTES     = 4   # Mapping Notes
COL_API_NAME  = 5   # SCRM Field Name (API)
COL_LABEL     = 6   # SCRM Field Label
COL_UAT_START = 7   # First UAT data column (G onwards)

HEADER_ROW   = 5   # Row containing column headers
DATA_START   = 6   # First data row
META_COL     = COL_LABEL  # Meta labels sit above SCRM Field Label (col F)


def _build_notes(existing: str, api_name: str, fields: dict,
                 source_systems: list) -> str:
    """
    Combine existing note with Andar Logic and/or DTracker Logic
    from the mapping template, separated by blank lines.
    """
    fd = fields.get(api_name, {})
    parts = []
    if existing and existing.strip():
        parts.append(existing.strip())
    if "Andar" in source_systems and fd.get("andar_logic"):
        parts.append(f"Andar Logic: {fd['andar_logic']}")
    if "DTracker" in source_systems and fd.get("dt_logic"):
        parts.append(f"DTracker Logic: {fd['dt_logic']}")
    return "\n\n".join(parts)


def build_output(
    blank_template_bytes: bytes,
    parsed: dict,
    source_systems: list,
    progress=None,
) -> bytes:
    """
    Build the complete output workbook.

    Args:
        blank_template_bytes : the hardcoded blank template as bytes
        parsed               : result of mapping_parser.parse_mapping_template()
        source_systems       : list of strings e.g. ["Andar", "DTracker"]
        progress             : optional callable(str) for status updates

    Returns:
        bytes of the finished .xlsx file
    """
    def log(msg):
        if progress:
            progress(msg)

    fields           = parsed["fields"]
    meta             = parsed["meta"]
    object_name      = meta["object_name"]
    api_names_ordered = meta["api_names_ordered"]

    log("Loading blank template…")
    wb = openpyxl.load_workbook(BytesIO(blank_template_bytes))
    ws = wb["Template"]

    # ── Capture header fill from the blank template before we overwrite ────────
    header_fill = copy(ws.cell(row=HEADER_ROW, column=1).fill)

    # ── Update meta rows (rows 1-4, col META_COL = col 5 = SCRM Field Label) ──
    log("Setting meta rows…")
    meta_labels = [
        "Tester",
        f"UAT {object_name}",
        "Original ID",
        f"UAT {object_name} ID",
    ]
    for i, label in enumerate(meta_labels, start=1):
        cell = ws.cell(row=i, column=META_COL, value=label)
        apply_std(cell, bold=True)

    # ── Update header row ──────────────────────────────────────────────────────
    log("Writing column headers…")

    # Determine database column headers based on selected source systems
    db_headers = []
    if "Andar" in source_systems:
        db_headers.append("Andar Location")
    if "DTracker" in source_systems:
        db_headers.append("DT Location")
    if "Raiser's Edge" in source_systems:
        db_headers.append("RE Location")

    # Pad to 3 database columns (A, B, C) with placeholders if fewer than 3
    placeholders = ["[Database 1] Location", "[Database 2] Location", "[Database 3] Location"]
    while len(db_headers) < 3:
        db_headers.append(placeholders[len(db_headers)])

    # Write A-C: database location headers
    for col_i, hdr in enumerate(db_headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_i, value=hdr)
        cell.fill = copy(header_fill)
        apply_std(cell, bold=True)

    # D=Mapping Notes, E=SCRM Field Name, F=SCRM Field Label
    fixed_headers = {
        COL_NOTES:    "Mapping Notes",
        COL_API_NAME: "SCRM Field Name",
        COL_LABEL:    "SCRM Field Label",
    }
    for col_i, hdr in fixed_headers.items():
        cell = ws.cell(row=HEADER_ROW, column=col_i, value=hdr)
        cell.fill = copy(header_fill)
        apply_std(cell, bold=True)

    # UAT data columns: replace [Object] with the detected object name
    for col_i in range(COL_UAT_START, ws.max_column + 1):
        cell = ws.cell(row=HEADER_ROW, column=col_i)
        if cell.value and "[Object]" in str(cell.value):
            n = col_i - COL_UAT_START + 1
            cell.value = f"{object_name} {n}"
        cell.fill = copy(header_fill)
        apply_std(cell, bold=True)

    # ── Write data rows ────────────────────────────────────────────────────────
    log(f"Writing {len(api_names_ordered)} field rows…")

    for row_offset, api_name in enumerate(api_names_ordered):
        row_num = DATA_START + row_offset
        fd = fields.get(api_name, {})

        # Col A: Andar Location
        andar_val = fd.get("andar_field", "") if "Andar" in source_systems else ""
        cell_a = ws.cell(row=row_num, column=COL_ANDAR, value=andar_val or None)
        apply_std(cell_a, wrap=bool(andar_val and "\n" in andar_val))

        # Col B: DT Location
        dt_val = fd.get("dt_field", "") if "DTracker" in source_systems else ""
        cell_b = ws.cell(row=row_num, column=COL_DT, value=dt_val or None)
        # Border rule: only if left (A) or right (C) neighbour is populated
        left_pop  = bool(andar_val)
        right_pop = bool(fd.get("andar_logic") or fd.get("dt_logic"))
        apply_std(cell_b,
                  wrap=bool(dt_val and "\n" in dt_val),
                  border=(left_pop or right_pop))

        # Col C: RE Location (if selected) — placeholder for now
        re_val = ""  # RE logic row detection can be added when a RE template exists
        cell_c = ws.cell(row=row_num, column=3, value=re_val or None)
        apply_std(cell_c)

        # Col D: Mapping Notes
        notes = _build_notes("", api_name, fields, source_systems)
        cell_d = ws.cell(row=row_num, column=COL_NOTES, value=notes or None)
        apply_std(cell_d, wrap=bool(notes and "\n" in notes))

        # Col E: SCRM Field Name (API name)
        cell_e = ws.cell(row=row_num, column=COL_API_NAME, value=api_name)
        apply_std(cell_e)

        # Col F: SCRM Field Label
        label = fd.get("scrm_label", "")
        cell_f = ws.cell(row=row_num, column=COL_LABEL, value=label or None)
        apply_std(cell_f)

        # UAT data columns G+: leave empty but apply std formatting
        for col_i in range(COL_UAT_START, ws.max_column + 1):
            apply_std(ws.cell(row=row_num, column=col_i))

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 40.0
    ws.column_dimensions["B"].width = 44.0
    ws.column_dimensions["C"].width = 40.0
    ws.column_dimensions["D"].width = 44.0
    ws.column_dimensions["E"].width = 30.0
    ws.column_dimensions["F"].width = 26.0

    # ── Build branch sheets ────────────────────────────────────────────────────
    log("Creating branch sheets…")

    for system in source_systems:
        for sheet_name, branch_type in BRANCH_CONFIGS.get(system, []):
            _build_branch(wb, ws, sheet_name, branch_type,
                          object_name, api_names_ordered, fields,
                          source_systems, header_fill)

    log("Saving…")
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _build_branch(wb, ws_template, sheet_name, branch_type,
                  object_name, api_names_ordered, fields,
                  source_systems, header_fill):
    """Create a single branch sheet."""
    # Remove existing sheet if present
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # Meta rows — col D (above SCRM Field Label)
    meta_labels = [
        "Tester",
        f"UAT {object_name}",
        "Original ID",
        f"UAT {object_name} ID",
    ]
    for i, label in enumerate(meta_labels, start=1):
        cell = ws.cell(row=i, column=META_COL, value=label)
        apply_std(cell, bold=True)

    # Header row
    loc_header = {
        "andar": "Andar Location",
        "dt":    "DT Location",
        "re":    "RE Location",
    }.get(branch_type, "Location")

    headers = [loc_header, "Mapping Notes", "SCRM Field Name", "SCRM Field Label"]
    uat_headers = [f"{object_name} {n}" for n in range(1, 26)]

    for col_i, hdr in enumerate(headers + uat_headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_i, value=hdr)
        cell.fill = copy(header_fill)
        apply_std(cell, bold=True)

    # Data rows
    for row_offset, api_name in enumerate(api_names_ordered):
        row_num = DATA_START + row_offset
        fd = fields.get(api_name, {})

        if branch_type == "dt":
            loc_val = fd.get("dt_field", "") if "DTracker" in source_systems else ""
        elif branch_type == "andar":
            loc_val = fd.get("andar_field", "") if "Andar" in source_systems else ""
        else:
            loc_val = ""

        notes = _build_notes("", api_name, fields, source_systems)
        label = fd.get("scrm_label", "")

        row_data = [loc_val or None, notes or None, api_name, label or None]
        for col_i, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_i, value=val)
            apply_std(cell, wrap=bool(val and "\n" in str(val)))

    # Column widths
    ws.column_dimensions["A"].width = 44.0
    ws.column_dimensions["B"].width = 44.0
    ws.column_dimensions["C"].width = 30.0
    ws.column_dimensions["D"].width = 26.0
