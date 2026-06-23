"""
transformer.py
Tab 1 logic: takes an existing UAT file + mapping template,
enriches it with DT Location / Mapping Notes, and creates branch sheets.
Accepts custom branch_overrides so sheet names are user-controlled.
"""

import re
from copy import copy
from io import BytesIO

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font

from mapping_parser import parse_mapping_template
from styles import apply_std, std_font, THIN_BORDER, NO_BORDER, swap_cols


def _detect_layout(ws) -> dict:
    """
    Auto-detect header row, data range, and column positions
    in an existing UAT sheet.
    """
    layout = {
        "header_row": None, "data_start": None, "data_end": None,
        "andar_col": None, "scrm_name_col": None, "scrm_label_col": None,
        "mapping_col": None, "uat_col_start": None, "uat_col_header": None,
        "meta_rows": [],
    }

    for row in ws.iter_rows(min_row=1, max_row=20):
        vals = [str(c.value or "").strip().lower() for c in row[:8]]
        joined = " ".join(vals)
        if any(kw in joined for kw in
               ["scrm field", "andar field", "andar location", "mapping notes"]):
            layout["header_row"] = row[0].row
            for cell in row:
                v = str(cell.value or "").strip().lower()
                if not v:
                    continue
                if "andar" in v and ("field" in v or "location" in v):
                    layout["andar_col"] = cell.column
                elif "scrm field name" in v:
                    layout["scrm_name_col"] = cell.column
                elif "scrm field label" in v:
                    layout["scrm_label_col"] = cell.column
                elif "mapping" in v:
                    layout["mapping_col"] = cell.column
                elif any(kw in v for kw in [
                    "payment", "relationship", "allocation", "gau",
                    "address", "affiliation", "federation"
                ]):
                    if layout["uat_col_start"] is None:
                        layout["uat_col_start"] = cell.column
                        layout["uat_col_header"] = str(cell.value or "").strip()
            break

    if layout["header_row"] is None:
        raise ValueError(
            "Could not detect the header row in the UAT file. "
            "Expected a row with 'Andar Field', 'SCRM Field Name', etc."
        )

    hr = layout["header_row"]
    for r in range(1, hr):
        for col_idx in range(1, 8):
            v = ws.cell(row=r, column=col_idx).value
            if v and str(v).strip():
                layout["meta_rows"].append((r, col_idx, str(v).strip()))
                break

    data_start = hr + 1
    data_end   = data_start
    for r in range(data_start, ws.max_row + 1):
        has_val = any(
            ws.cell(row=r, column=c).value not in (None, "")
            for c in range(1, 6)
        )
        if has_val:
            data_end = r
        else:
            next_has = any(
                ws.cell(row=r + 1, column=c).value not in (None, "")
                for c in range(1, 6)
            )
            if not next_has:
                break

    layout["data_start"] = data_start
    layout["data_end"]   = data_end
    return layout


def _build_notes(existing, api_name, fields, source_systems):
    fd = fields.get(api_name, {})
    parts = []
    if existing and str(existing).strip():
        parts.append(str(existing).strip())
    if "Andar" in source_systems and fd.get("andar_logic"):
        parts.append(f"Andar Logic: {fd['andar_logic']}")
    if "DTracker" in source_systems and fd.get("dt_logic"):
        parts.append(f"DTracker Logic: {fd['dt_logic']}")
    return "\n\n".join(parts)


def transform(
    uat_bytes: bytes,
    template_bytes: bytes,
    source_systems: list,
    branch_overrides: list = None,
    progress_callback=None,
) -> bytes:
    """
    Enrich an existing UAT file using the mapping template.

    Args:
        uat_bytes        : the existing UAT tracking file
        template_bytes   : the strategy mapping template
        source_systems   : e.g. ["Andar", "DTracker"]
        branch_overrides : list of (sheet_name, branch_type) for custom naming
        progress_callback: optional callable(str)
    """
    def log(msg):
        if progress_callback: progress_callback(msg)

    log("Parsing mapping template…")
    parsed = parse_mapping_template(template_bytes)
    fields = parsed["fields"]

    log("Loading UAT file…")
    wb = openpyxl.load_workbook(BytesIO(uat_bytes))

    # Rename Sheet1 → Template
    for name in list(wb.sheetnames):
        if name.strip().lower() in ("sheet1", "sheet 1"):
            wb[name].title = "Template"
            break

    if "Template" not in wb.sheetnames:
        raise ValueError(
            "Could not find a sheet named 'Template', 'Sheet1', or 'Sheet 1'."
        )

    log("Analysing UAT layout…")
    ws_t   = wb["Template"]
    layout = _detect_layout(ws_t)
    hr     = layout["header_row"]
    ds     = layout["data_start"]
    de     = layout["data_end"]
    scrm_lookup_col = layout["scrm_name_col"] or layout["andar_col"]

    # ── Ensure Andar col is col 1 ──────────────────────────────────────────────
    if layout["andar_col"] and layout["andar_col"] != 1:
        swap_cols(ws_t, 1, layout["andar_col"])
        old_a = layout["andar_col"]
        for key in ("scrm_name_col", "scrm_label_col", "mapping_col", "uat_col_start"):
            if layout[key] == 1:
                layout[key] = old_a
        layout["andar_col"] = 1

    # ── Move meta rows to sit above SCRM Field Label after insert ──────────────
    scrm_label_col = layout["scrm_label_col"]
    if scrm_label_col and scrm_label_col > 1:
        target = scrm_label_col - 1
        for (r, cur_col, val) in layout["meta_rows"]:
            if cur_col != target:
                ws_t.cell(row=r, column=target).value = val
                ws_t.cell(row=r, column=cur_col).value = None

    # ── Insert DT Location column (col B) ─────────────────────────────────────
    ws_t.insert_cols(2)
    for key in ("scrm_name_col", "scrm_label_col", "mapping_col", "uat_col_start"):
        if layout[key] and layout[key] >= 2:
            layout[key] += 1
    if scrm_lookup_col and scrm_lookup_col >= 2:
        scrm_lookup_col += 1

    # ── Ensure Mapping Notes is col 3 ─────────────────────────────────────────
    if layout["mapping_col"] and layout["mapping_col"] != 3:
        mc = layout["mapping_col"]
        direction = -1 if mc > 3 else 1
        while mc != 3:
            swap_cols(ws_t, mc, mc + direction)
            mc += direction
        layout["mapping_col"] = 3
        for cell in ws_t[hr]:
            v = str(cell.value or "").strip().lower()
            if "scrm field name" in v:
                layout["scrm_name_col"] = cell.column
                scrm_lookup_col = cell.column
            elif "scrm field label" in v:
                layout["scrm_label_col"] = cell.column

    # ── Update headers ─────────────────────────────────────────────────────────
    log("Updating headers…")
    ws_t.cell(row=hr, column=1).value = "Andar Location"
    ws_t.cell(row=hr, column=2).value = "DT Location"
    if not ws_t.cell(row=hr, column=3).value:
        ws_t.cell(row=hr, column=3).value = "Mapping Notes"
    ws_t.cell(row=hr, column=2).fill = copy(ws_t.cell(row=hr, column=1).fill)
    for col in range(1, ws_t.max_column + 1):
        c = ws_t.cell(row=hr, column=col)
        if c.value:
            apply_std(c, bold=True)

    # Style meta rows
    new_meta_col = layout["scrm_label_col"] or 5
    for r in range(1, hr):
        c = ws_t.cell(row=r, column=new_meta_col)
        if c.value:
            apply_std(c, bold=True)

    # ── Populate data rows ─────────────────────────────────────────────────────
    log("Populating DT Location and Mapping Notes…")
    for row_num in range(ds, de + 1):
        api_name = str(
            ws_t.cell(row=row_num, column=scrm_lookup_col).value or ""
        ).strip()
        fd = fields.get(api_name, {})

        dt_val = fd.get("dt_field", "") if "DTracker" in source_systems else ""
        ws_t.cell(row=row_num, column=2).value = dt_val or None

        existing = ws_t.cell(row=row_num, column=3).value or ""
        notes    = _build_notes(existing, api_name, fields, source_systems)
        if notes:
            ws_t.cell(row=row_num, column=3).value = notes

        if "Andar" in source_systems and fd.get("andar_field"):
            ws_t.cell(row=row_num, column=1).value = fd["andar_field"]

    # ── Apply formatting ───────────────────────────────────────────────────────
    log("Applying formatting…")
    for row in ws_t.iter_rows(min_row=ds, max_row=de):
        for cell in row:
            apply_std(cell,
                      wrap=bool(cell.value and "\n" in str(cell.value)))

    # DT border rule
    for r in range(ds, de + 1):
        left  = ws_t.cell(row=r, column=1).value
        right = ws_t.cell(row=r, column=3).value
        cell_b = ws_t.cell(row=r, column=2)
        if left or right:
            apply_std(cell_b,
                      wrap=bool(cell_b.value and "\n" in str(cell_b.value or "")))
        else:
            cell_b.border    = Border()
            cell_b.font      = std_font()
            cell_b.alignment = Alignment()

    ws_t.column_dimensions["A"].width = 40.0
    ws_t.column_dimensions["B"].width = 44.0
    ws_t.column_dimensions["C"].width = 44.0
    ws_t.column_dimensions["D"].width = 28.0
    ws_t.column_dimensions["E"].width = 28.0

    # ── Branch sheets ──────────────────────────────────────────────────────────
    log("Creating branch sheets…")
    if branch_overrides is None:
        defaults = {
            "Andar":         [("WHALIF", "andar")],
            "DTracker":      [("WFREDE", "dt")],
            "Raiser's Edge": [("WRE1",   "re")],
        }
        branch_overrides = [
            pair for sys in source_systems
            for pair in defaults.get(sys, [])
        ]

    header_fill = copy(ws_t.cell(row=hr, column=1).fill)
    object_name = parsed["meta"]["object_name"]
    api_names   = parsed["meta"]["api_names_ordered"]

    for sheet_name, branch_type in branch_overrides:
        if not sheet_name:
            continue
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws_b = wb.create_sheet(sheet_name)

        # Meta
        for i, label in enumerate(
            ["Tester", f"UAT {object_name}", "Original ID",
             f"UAT {object_name} ID"], start=1
        ):
            apply_std(ws_b.cell(row=i, column=new_meta_col, value=label), bold=True)

        # Header
        loc_hdr = {"andar": "Andar Location",
                   "dt":    "DT Location",
                   "re":    "RE Location"}.get(branch_type, "Location")
        uat_prefix = re.sub(r"\s*\d+$", "",
            str(ws_t.cell(row=hr, column=layout.get("uat_col_start") or 6
                          ).value or object_name)
        ).strip() or object_name

        all_hdrs = ([loc_hdr, "Mapping Notes", "SCRM Field Name", "SCRM Field Label"] +
                    [f"{uat_prefix} {n}" for n in range(1, 26)])
        for col_i, hdr in enumerate(all_hdrs, start=1):
            cell = ws_b.cell(row=hr, column=col_i, value=hdr)
            cell.fill = copy(header_fill)
            apply_std(cell, bold=True)

        # Data
        for src_row in range(ds, de + 1):
            andar_v = ws_t.cell(row=src_row, column=1).value
            dt_v    = ws_t.cell(row=src_row, column=2).value
            note_v  = ws_t.cell(row=src_row, column=3).value
            name_v  = ws_t.cell(row=src_row,
                                 column=layout["scrm_name_col"] or 4).value
            label_v = ws_t.cell(row=src_row,
                                 column=layout["scrm_label_col"] or 5).value

            loc_val = dt_v if branch_type == "dt" else andar_v
            for col_i, val in enumerate(
                [loc_val, note_v, name_v, label_v], start=1
            ):
                apply_std(ws_b.cell(row=src_row, column=col_i, value=val),
                          wrap=bool(val and "\n" in str(val)))

        for col_i, w in enumerate([44.0, 44.0, 28.0, 26.0], start=1):
            from openpyxl.utils import get_column_letter
            ws_b.column_dimensions[get_column_letter(col_i)].width = w

    # ── Update existing tester sheets ──────────────────────────────────────────
    skip = {"Template", "Legend", "MAPPING", "Mapping"} | {
        n for n, _ in branch_overrides
    }
    for sname in wb.sheetnames:
        if sname in skip:
            continue
        ws_b = wb[sname]
        try:
            blayout = _detect_layout(ws_b)
            if blayout["header_row"] is None:
                continue
            bhr = blayout["header_row"]
            bds = blayout["data_start"]
            bde = blayout["data_end"]
            has_dt = any(
                "dt location" in str(ws_b.cell(row=bhr, column=c).value or "").lower()
                for c in range(1, 8)
            )
            if not has_dt:
                ws_b.insert_cols(2)
                ws_b.cell(row=bhr, column=2).value = "DT Location"
                ws_b.cell(row=bhr, column=2).fill = copy(
                    ws_b.cell(row=bhr, column=1).fill
                )
            ws_b.cell(row=bhr, column=1).value = "Andar Location"
            for col in range(1, ws_b.max_column + 1):
                c = ws_b.cell(row=bhr, column=col)
                if c.value:
                    apply_std(c, bold=True)
            for src_row in range(bds, bde + 1):
                dt_val   = ws_t.cell(row=src_row, column=2).value
                note_val = ws_t.cell(row=src_row, column=3).value
                apply_std(ws_b.cell(row=src_row, column=2, value=dt_val),
                          wrap=bool(dt_val and "\n" in str(dt_val or "")))
                apply_std(ws_b.cell(row=src_row, column=3, value=note_val),
                          wrap=bool(note_val and "\n" in str(note_val or "")))
            for row in ws_b.iter_rows(min_row=bds, max_row=bde):
                for cell in row:
                    if cell.column not in (2, 3) and cell.value is not None:
                        apply_std(cell,
                                  wrap=bool(cell.value and "\n" in str(cell.value)))
        except Exception:
            pass

    log("Saving…")
    out = BytesIO()
    wb.save(out)
    return out.getvalue()
